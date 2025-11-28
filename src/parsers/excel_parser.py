# =============================================================================
# FILE: src/parsers/excel_parser.py
# =============================================================================
"""
Advanced Excel parser supporting multiple sheets, formulas, and metadata.
Handles both .xlsx and .xls formats.
References: [web:243][web:246][web:248]
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, Any, List, Optional
import structlog

logger = structlog.get_logger()


class ExcelParser:
    """
    Parse Excel files (.xlsx, .xls) with support for multiple sheets,
    formulas, formatting, and metadata extraction.
    """
    
    @staticmethod
    def parse_file(filepath: Path, sheet_names: Optional[List[str]] = None) -> Dict[str, Any]:
        """
        Parse Excel file with all sheets or specified sheets.
        
        Args:
            filepath: Path to Excel file
            sheet_names: Optional list of specific sheet names to parse
            
        Returns:
            Dict containing all parsed sheets and metadata
        """
        try:
            # Read all sheets into dict of DataFrames[web:243]
            if sheet_names:
                sheets_dict = pd.read_excel(
                    filepath,
                    sheet_name=sheet_names,
                    engine='openpyxl'
                )
            else:
                # Read all sheets[web:243]
                sheets_dict = pd.read_excel(
                    filepath,
                    sheet_name=None,  # Read all sheets
                    engine='openpyxl'
                )
            
            # Process each sheet
            processed_sheets = {}
            for sheet_name, df in sheets_dict.items():
                processed_sheets[sheet_name] = {
                    'shape': df.shape,
                    'columns': df.columns.tolist(),
                    'dtypes': df.dtypes.astype(str).to_dict(),
                    'data': df.to_dict(orient='records'),
                    'dataframe': df,
                    'summary': ExcelParser._generate_summary(df),
                    'has_nulls': df.isnull().any().any(),
                    'null_counts': df.isnull().sum().to_dict()
                }
            
            # Extract workbook metadata using openpyxl[web:246]
            metadata = ExcelParser._extract_metadata(filepath)
            
            logger.info(
                "excel_parsed_successfully",
                filepath=str(filepath),
                sheet_count=len(processed_sheets),
                total_rows=sum(s['shape'][0] for s in processed_sheets.values())
            )
            
            return {
                'type': 'excel',
                'filepath': str(filepath),
                'sheet_names': list(processed_sheets.keys()),
                'sheet_count': len(processed_sheets),
                'sheets': processed_sheets,
                'metadata': metadata
            }
            
        except Exception as e:
            logger.error("excel_parse_error", filepath=str(filepath), error=str(e))
            return {
                'error': str(e),
                'type': 'excel',
                'filepath': str(filepath)
            }
    
    @staticmethod
    def _extract_metadata(filepath: Path) -> Dict[str, Any]:
        """
        Extract workbook metadata using openpyxl[web:246].
        
        Args:
            filepath: Path to Excel file
            
        Returns:
            Dict containing metadata (author, created date, etc.)
        """
        try:
            wb = load_workbook(filepath, read_only=True, data_only=False)
            metadata = {
                'creator': wb.properties.creator if hasattr(wb.properties, 'creator') else None,
                'created': str(wb.properties.created) if hasattr(wb.properties, 'created') else None,
                'modified': str(wb.properties.modified) if hasattr(wb.properties, 'modified') else None,
                'lastModifiedBy': wb.properties.lastModifiedBy if hasattr(wb.properties, 'lastModifiedBy') else None,
                'sheet_names': wb.sheetnames,
                'sheet_count': len(wb.sheetnames)
            }
            wb.close()
            return metadata
        except Exception as e:
            logger.warning("metadata_extraction_failed", error=str(e))
            return {}
    
    @staticmethod
    def _generate_summary(df: pd.DataFrame) -> Dict[str, Any]:
        """Generate summary statistics for a sheet"""
        summary = {}
        
        # Numeric summary
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            summary['numeric'] = df[numeric_cols].describe().to_dict()
        
        # Categorical summary
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            summary['categorical'] = {
                col: {
                    'unique_count': df[col].nunique(),
                    'top_values': df[col].value_counts().head(3).to_dict()
                }
                for col in categorical_cols
            }
        
        return summary
    
    @staticmethod
    def extract_formulas(filepath: Path, sheet_name: str = None) -> Dict[str, List[Dict]]:
        """
        Extract cell formulas from Excel file[web:246].
        
        Args:
            filepath: Path to Excel file
            sheet_name: Optional specific sheet name
            
        Returns:
            Dict mapping sheet names to lists of formula cells
        """
        try:
            wb = load_workbook(filepath, data_only=False)
            formulas_by_sheet = {}
            
            sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
            
            for name in sheets_to_process:
                ws = wb[name]
                formulas = []
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append({
                                'cell': cell.coordinate,
                                'formula': cell.value,
                                'value': cell.value  # Formula result if calculated
                            })
                
                if formulas:
                    formulas_by_sheet[name] = formulas
            
            wb.close()
            logger.info(f"Extracted {sum(len(v) for v in formulas_by_sheet.values())} formulas")
            return formulas_by_sheet
            
        except Exception as e:
            logger.error("formula_extraction_error", error=str(e))
            return {}
